import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# إعداد الصفحة
st.set_page_config(page_title="GWA MASR | Travel Dashboard", layout="wide", page_icon="🌍")

# 🎨 تنسيق CSS احترافي جدًا لشركة سياحة
st.markdown("""
<style>
body {
    background: linear-gradient(135deg, #001F3F 0%, #003B73 50%, #0077B6 100%);
    color: white;
    font-family: 'Poppins', sans-serif;
}
h1 {
    font-size: 3.2em;
    font-weight: 800;
    color: #FFD700;
    text-align: center;
    text-shadow: 0 0 20px rgba(255,215,0,0.7), 0 0 40px #00B4D8;
    letter-spacing: 2px;
    margin-bottom: 0.2em;
}
h2 {
    color: #ADE8F4;
    text-align: center;
    font-weight: 400;
    margin-top: -10px;
    letter-spacing: 1px;
}
.stTabs [data-baseweb="tab-list"] {
    justify-content: center;
}
.stTabs [data-baseweb="tab"] {
    background-color: #023E8A;
    color: white;
    border-radius: 10px 10px 0 0;
    padding: 10px 20px;
    font-weight: bold;
    transition: all 0.3s ease;
}
.stTabs [data-baseweb="tab"]:hover {
    background-color: #0077B6;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(90deg, #0077B6, #00B4D8);
    color: #FFD700 !important;
    font-weight: bold;
    transform: scale(1.05);
}
.stButton>button {
    background: linear-gradient(90deg, #0077B6, #00B4D8);
    color: white;
    border: none;
    border-radius: 10px;
    font-weight: bold;
    padding: 0.6em 1.4em;
    transition: 0.3s ease;
}
.stButton>button:hover {
    transform: scale(1.05);
    box-shadow: 0 0 20px rgba(0,180,216,0.6);
}
[data-testid="stDataFrame"] {
    border: 1px solid #00B4D8;
    border-radius: 10px;
}
.footer {
    text-align: center;
    font-size: 0.9em;
    color: #BBBBBB;
    margin-top: 40px;
}
.glow {
    animation: glow 2s ease-in-out infinite alternate;
}
@keyframes glow {
    from { text-shadow: 0 0 10px #FFD700, 0 0 20px #FFD700; }
    to { text-shadow: 0 0 20px #00B4D8, 0 0 40px #FFD700; }
}
</style>
""", unsafe_allow_html=True)

# 🌍 رأس الصفحة
st.markdown(
    """
    <h1 style='text-align: center; color: #1E90FF;'>
        ✈️ <b>GWA MASR | Travel Dashboard</b> 🌍
    </h1>
    """,
    unsafe_allow_html=True
)




# 📊 رابط الشيت
sheet_url = "https://docs.google.com/spreadsheets/d/1pu0_BsqhJcQfsPo6R7AckvsnHhvMit6HmHH-dQDtO2s/gviz/tq?tqx=out:csv"

try:
    df = pd.read_csv(sheet_url)
    st.success("✅ Data loaded successfully!")

    df.columns = df.columns.str.strip()

    # 🔹 تحويل الأعمدة اللي فيها تاريخ إلى تاريخ فقط بدون وقت
    for col in df.columns:
        if any(word in col.lower() for word in ["date", "check in", "check out", "check"]):
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
            except:
                pass

    # 🔹 تحويل أرقام الهواتف
    for col in df.columns:
        if any(word in col.lower() for word in ["phone", "mobile", "number"]):
            df[col] = df[col].astype(str).str.replace(".0", "", regex=False)

    # 🔹 تحويل الكراسي لأرقام صحيحة
    for col in df.columns:
        if any(word in col.lower() for word in ["seat", "chair", "transfer"]):
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    tab_home, tab1, tab2, tab3, tab4 , tab5 = st.tabs([
    "🏠 Home Overview", 
    "📊 General Filter", 
    "🚌 Situation Dahab", 
    " situation siwa 🚌 ", 
    " 🛏️ rooming list " ,
    "Avalability 🏨"
    ])


    with tab_home:
        st.markdown("## 🏠 Home Overview — Daily Destinations Snapshot")

        # محاولة اكتشاف أعمدة مهمة تلقائياً
        checkin_col = next((c for c in df.columns if "check" in c.lower() and "in" in c.lower()), None)
        dest_col = next((c for c in df.columns if "dest" in c.lower() or "destination" in c.lower()), None)
        seats_col = next((c for c in df.columns if any(w in c.lower() for w in ["seat", "no of seats", "no.seats", "seats", "no of seat"])), None)
        from_col = next((c for c in df.columns if "from" in c.lower()), None)
        nights_col = next((c for c in df.columns if "night" in c.lower()), None)
        unpaid_col = next((c for c in df.columns if any(w in c.lower() for w in ["un paid", "unpaid", "balance", "due"])), None)
        name_col = next((c for c in df.columns if "name" in c.lower()), None)
        phone_col = next((c for c in df.columns if any(w in c.lower() for w in ["phone", "mobile", "tel", "number"])), None)
        company_col = next((c for c in df.columns if any(w in c.lower() for w in ["company", "comp", "agency", "office"])), None)

        if not checkin_col or not dest_col:
            st.error("❌ لا أجد أعمدة 'check in' أو 'dest' في الملف — تأكد من أسماء الأعمدة في الشيت.")
        else:
            df[checkin_col] = pd.to_datetime(df[checkin_col], errors='coerce').dt.date
            check_in_dates = sorted(df[checkin_col].dropna().unique())
            selected_date = st.selectbox("📅 اختر تاريخ Check In لعرض الـ Overview", check_in_dates)

            if selected_date:
                day_df = df[df[checkin_col] == selected_date].copy()

                if day_df.empty:
                    st.warning("⚠️ لا توجد حجوزات في التاريخ المحدد.")
                else:
                    st.success(f"✅ {len(day_df)} حجز سيتم الدخول يوم {selected_date}")

                    # تنظيف الأعمدة
                    if seats_col:
                        day_df[seats_col] = pd.to_numeric(day_df[seats_col], errors='coerce').fillna(0).astype(int)
                    if nights_col:
                        day_df[nights_col] = day_df[nights_col].astype(str).str.extract(r'(\d+)')[0].fillna("0")
                    if from_col:
                        day_df[from_col] = day_df[from_col].astype(str).str.title().fillna("Unknown")
                    if unpaid_col:
                        day_df[unpaid_col] = pd.to_numeric(day_df[unpaid_col].astype(str).str.replace(r'[^\d\.-]', '', regex=True), errors='coerce').fillna(0)

                    # تجميع حسب الوجهة
                    group = day_df.groupby(day_df[dest_col].astype(str).str.title(), dropna=True)

                    cols_per_row = 3
                    dest_list = list(group.groups.keys())

                    def make_excel_from_df(dframe, sheet_name="Sheet1"):
                        wb = Workbook()
                        ws = wb.active
                        ws.title = sheet_name
                        for ci, colname in enumerate(dframe.columns, start=1):
                            ws.cell(row=1, column=ci, value=colname)
                        for ri, row in enumerate(dframe.values, start=2):
                            for ci, v in enumerate(row, start=1):
                                ws.cell(row=ri, column=ci, value=str(v))
                        buf = BytesIO()
                        wb.save(buf)
                        buf.seek(0)
                        return buf

                    # Download كامل اليوم
                    overall_buf = make_excel_from_df(day_df, sheet_name=f"Overview_{selected_date}")
                    st.download_button(
                        label="📘 Download Full Day Overview (Excel)",
                        data=overall_buf,
                        file_name=f"Overview_{selected_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # عرض الوجهات في كروت
                    rows = (len(dest_list) + cols_per_row - 1) // cols_per_row
                    idx = 0
                    for r in range(rows):
                        cols = st.columns(cols_per_row)
                        for c in cols:
                            if idx >= len(dest_list):
                                c.write("")
                                idx += 1
                                continue

                            dest = dest_list[idx]
                            dest_df = group.get_group(dest)
                            idx += 1

                            total_bookings = len(dest_df)
                            unpaid_total = float(dest_df[unpaid_col].sum()) if unpaid_col else 0.0

                            # seats & nights grouping
                            seats_from = {}
                            nights_from = {}
                            if seats_col and from_col:
                                for f, g in dest_df.groupby(from_col):
                                    seats_from[f] = int(g[seats_col].sum())
                                    if nights_col:
                                        nights_from[f] = g[nights_col].value_counts().to_dict()

                            # box HTML
                            box_content = ""
                            for f, seats in seats_from.items():
                                box_content += f"<b>{f}:</b> {seats} seats<br>"
                                if nights_col and f in nights_from:
                                    for night, cnt in sorted(nights_from[f].items(), reverse=True):
                                        box_content += f"&nbsp;&nbsp;{night} nights : {cnt}<br>"

                            box_html = f"""
                            <div style="border-radius:12px; padding:14px; margin:8px; background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(0,0,0,0.06)); border:1px solid rgba(255,255,255,0.06)">
                                <h3 style="margin:0; color:#FFD700;">📍 {dest} 
                                    <small style="color:#ADE8F4; font-weight:400; font-size:0.8em">
                                        ({total_bookings} bookings)
                                    </small>
                                </h3>
                                <p style="margin:6px 0 8px 0; color:#E6F7FF;">
                                    <b>Unpaid Total:</b> {unpaid_total:,.2f}
                                </p>
                                <p style="margin:4px 0; color:#DDEFFD;">{box_content}</p>
                            </div>
                            """
                            c.markdown(box_html, unsafe_allow_html=True)

                            # Expander التفاصيل
                            with c.expander(f"🔎 تفاصيل {dest} ({total_bookings}):"):

                                possible_cols = [
                                    name_col, phone_col, company_col,
                                    seats_col, nights_col, unpaid_col,
                                    from_col, checkin_col
                                ]

                                detail_cols = [
                                    col for col in possible_cols
                                    if col is not None and col in dest_df.columns
                                ]

                                st.dataframe(dest_df[detail_cols].reset_index(drop=True), use_container_width=True)

                                buf_dest = make_excel_from_df(dest_df.reset_index(drop=True),
                                                            sheet_name=f"{dest}_{selected_date}")
                                st.download_button(
                                    label=f"📥 Download {dest} Details (Excel)",
                                    data=buf_dest,
                                    file_name=f"{dest.replace(' ','_')}_Details_{selected_date}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

    # ✅---------------------- TAB 1 --------------------------
    with tab1:
        with st.expander("📋 Preview Data"):
            st.dataframe(df.head(), use_container_width=True)

        st.markdown("### 🎯 Step 1: Choose columns to filter by")
        filter_cols = st.multiselect("Select columns to filter", df.columns.tolist())

        filters = {}
        if filter_cols:
            st.markdown("### 🧩 Step 2: Choose filter values")
            for col in filter_cols:
                unique_vals = sorted(df[col].dropna().astype(str).unique())
                selected_vals = st.multiselect(f"Select values for **{col}**", ["All"] + unique_vals, default=["All"])
                if "All" not in selected_vals:
                    filters[col] = selected_vals

        filtered_df = df.copy()
        for col, vals in filters.items():
            filtered_df = filtered_df[filtered_df[col].astype(str).isin(vals)]

        st.markdown("### 🧾 Step 3: Choose columns to display")
        selected_features = st.multiselect("Select columns to display", df.columns.tolist(), default=df.columns[:5])

        if st.button("🔍 Show Filtered Data"):
            if filtered_df.empty:
                st.warning("⚠️ No records found for the selected filters.")
            else:
                # ✅ تحويل التاريخ لتاريخ فقط
                for col in filtered_df.columns:
                    if any(word in col.lower() for word in ["check in", "check out", "date", "check"]):
                        try:
                            filtered_df[col] = pd.to_datetime(filtered_df[col], errors='coerce').dt.date
                        except:
                            pass

                st.success(f"✅ Showing {len(filtered_df)} matching rows")
                st.dataframe(filtered_df[selected_features], use_container_width=True)

                # 🔷 إنشاء ملف Excel منسق
                wb = Workbook()
                ws = wb.active
                ws.title = "Filtered Data"

                header_fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                align_center = Alignment(horizontal="center", vertical="center")
                border = Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC")
                )

                for col_num, column_title in enumerate(selected_features, start=1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = border

                for row_num, row_data in enumerate(filtered_df[selected_features].values, start=2):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_num, column=col_num, value=str(cell_value))
                        if row_num % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.alignment = align_center
                        cell.border = border

                for col in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in col if cell.value)
                    ws.column_dimensions[col[0].column_letter].width = max_length + 2

                footer_row = len(filtered_df) + 3
                ws.merge_cells(f"A{footer_row}:{ws.cell(1, len(selected_features)).column_letter}{footer_row}")
                ws.cell(row=footer_row, column=1).value = f"Generated by GWA MASR | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                ws.cell(row=footer_row, column=1).alignment = align_center
                ws.cell(row=footer_row, column=1).font = Font(color="888888", italic=True, size=10)

                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                st.download_button(
                    label="📘 Download Filtered Data (Excel)",
                    data=buffer,
                    file_name="Filtered_Bookings.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # ✅---------------------- TAB 2 --------------------------
    with tab2:
        st.markdown("## 🏝️ Situation Dahab (Bus Trips Only)")

        checkin_col = next((c for c in df.columns if "check" in c.lower() and "in" in c.lower()), None)
        checkout_col = next((c for c in df.columns if "check" in c.lower() and "out" in c.lower()), None)
        dest_col = next((c for c in df.columns if "dest" in c.lower()), None)
        type_col = next((c for c in df.columns if "type" in c.lower()), None)

        if checkin_col in df.columns:
            df[checkin_col] = pd.to_datetime(df[checkin_col], errors='coerce').dt.date
        if checkout_col in df.columns:
            df[checkout_col] = pd.to_datetime(df[checkout_col], errors='coerce').dt.date

        if not all([checkin_col, dest_col, type_col]):
            st.error("❌ One or more required columns not found in your sheet.")
        else:
            check_in_dates = sorted(df[checkin_col].dropna().unique())
            selected_date = st.selectbox("📅 Select Check In Date", check_in_dates)

            if selected_date:
                filtered_dahab = df[
                    (df[dest_col].astype(str).str.lower() == "dahab") &
                    (df[type_col].astype(str).str.lower() == "bus") &
                    (df[checkin_col] == selected_date)
                ]

                if filtered_dahab.empty:
                    st.warning("⚠️ No Dahab Bus trips found for this date.")
                else:
                    st.success(f"✅ Found {len(filtered_dahab)} bookings for Dahab Bus on {selected_date}")

                    cols_to_show = [
                        "name", "accomadation phone number", "company", "hotels", "Rooms",
                        "No of seats", "check in H", "check out H", "No.nights", "Un paid", "from"
                    ]
                    existing_cols = [c for c in cols_to_show if c in filtered_dahab.columns]
                    st.dataframe(filtered_dahab[existing_cols], use_container_width=True)

                    # 🔷 نفس تنسيق Excel الأزرق
                    wb2 = Workbook()
                    ws2 = wb2.active
                    ws2.title = "Situation Dahab"

                    header_fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    align_center = Alignment(horizontal="center", vertical="center")
                    border = Border(
                        left=Side(style="thin", color="CCCCCC"),
                        right=Side(style="thin", color="CCCCCC"),
                        top=Side(style="thin", color="CCCCCC"),
                        bottom=Side(style="thin", color="CCCCCC")
                    )

                    for col_num, column_title in enumerate(existing_cols, start=1):
                        cell = ws2.cell(row=1, column=col_num, value=column_title)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = align_center
                        cell.border = border

                    for row_num, row_data in enumerate(filtered_dahab[existing_cols].values, start=2):
                        for col_num, cell_value in enumerate(row_data, start=1):
                            cell = ws2.cell(row=row_num, column=col_num, value=str(cell_value))
                            if row_num % 2 == 0:
                                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                            cell.alignment = align_center
                            cell.border = border

                    for col in ws2.columns:
                        max_length = max(len(str(cell.value)) for cell in col if cell.value)
                        ws2.column_dimensions[col[0].column_letter].width = max_length + 2

                    footer_row = len(filtered_dahab) + 3
                    ws2.merge_cells(f"A{footer_row}:{ws2.cell(1, len(existing_cols)).column_letter}{footer_row}")
                    ws2.cell(row=footer_row, column=1).value = f"Generated by GWA MASR | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    ws2.cell(row=footer_row, column=1).alignment = align_center
                    ws2.cell(row=footer_row, column=1).font = Font(color="888888", italic=True, size=10)

                    buffer2 = BytesIO()
                    wb2.save(buffer2)
                    buffer2.seek(0)

                    st.download_button(
                        label="📘 Download Situation Dahab (Excel)",
                        data=buffer2,
                        file_name=f"Situation_Dahab_{selected_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        # ✅---------------------- TAB 3: Situation Siwa --------------------------
    with tab3:
        st.markdown("## 🏜️ Situation Siwa (Bus Trips Only)")

        # محاولة تحديد الأعمدة تلقائيًا
        checkin_col = next((c for c in df.columns if "check" in c.lower() and "in" in c.lower()), None)
        checkout_col = next((c for c in df.columns if "check" in c.lower() and "out" in c.lower()), None)
        dest_col = next((c for c in df.columns if "dest" in c.lower()), None)
        type_col = next((c for c in df.columns if "type" in c.lower()), None)
        from_col = next((c for c in df.columns if "from" in c.lower()), None)
        nights_col = next((c for c in df.columns if "night" in c.lower()), None)

        # تحويل الأعمدة لتاريخ فقط بدون وقت
        if checkin_col in df.columns:
            df[checkin_col] = pd.to_datetime(df[checkin_col], errors='coerce').dt.date
        if checkout_col in df.columns:
            df[checkout_col] = pd.to_datetime(df[checkout_col], errors='coerce').dt.date

        if not all([checkin_col, dest_col, type_col, from_col, nights_col]):
            st.error("❌ One or more required columns not found in your sheet.")
        else:
            df[from_col] = df[from_col].astype(str).str.lower().str.strip()
            df[nights_col] = df[nights_col].astype(str).str.extract(r'(\d+)')[0]

            # ✅ 1️⃣ اختيار يوم Check In
            check_in_dates = sorted(df[checkin_col].dropna().unique())
            selected_date = st.selectbox("📅 Select Check In Date", check_in_dates, key="siwa_checkin")

            if selected_date:
                filtered_step1 = df[
                    (df[checkin_col] == selected_date) &
                    (df[dest_col].astype(str).str.lower().str.contains("siwa")) &
                    (df[type_col].astype(str).str.lower().str.contains("bus"))
                ]

                if filtered_step1.empty:
                    st.warning("⚠️ No Siwa trips found for this Check In date.")
                else:
                    # ✅ 2️⃣ اختيار From (Alex / Cairo)
                    from_options = sorted(filtered_step1[from_col].dropna().unique())
                    selected_from = st.selectbox("📍 Select 'From' location", from_options, key="siwa_from")

                    filtered_step2 = filtered_step1[filtered_step1[from_col] == selected_from]

                    if selected_from:
                        # ✅ 3️⃣ اختيار عدد الليالي بناءً على الـ From
                        if selected_from.lower() == "alex":
                            allowed_nights = ["2", "3"]
                        elif selected_from.lower() == "cairo":
                            allowed_nights = ["3"]
                        else:
                            allowed_nights = sorted(filtered_step2[nights_col].dropna().unique())

                        nights_available = [n for n in allowed_nights if n in filtered_step2[nights_col].astype(str).unique()]
                        selected_nights = st.selectbox("🌙 Select No. of Nights", nights_available, key="siwa_nights")

                        final_siwa = filtered_step2[filtered_step2[nights_col] == selected_nights]

                        if final_siwa.empty:
                            st.warning("⚠️ No Siwa Bus trips found with these filters.")
                        else:
                            st.success(f"✅ Found {len(final_siwa)} Siwa Bus bookings on {selected_date} from {selected_from.title()} ({selected_nights} nights)")

                            cols_to_show = [
                                "name", "accomadation phone number", "company", "hotels", "Rooms",
                                "No of seats", "check in H", "check out H", 
                                "No.nights", "Un paid", "from"
                            ]
                            existing_cols = [c for c in cols_to_show if c in final_siwa.columns]
                            st.dataframe(final_siwa[existing_cols], use_container_width=True)

                            # 🔷 إنشاء Excel بنفس التنسيق الأزرق
                            wb3 = Workbook()
                            ws3 = wb3.active
                            ws3.title = "Situation Siwa"

                            header_fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
                            header_font = Font(color="FFFFFF", bold=True)
                            align_center = Alignment(horizontal="center", vertical="center")
                            border = Border(
                                left=Side(style="thin", color="CCCCCC"),
                                right=Side(style="thin", color="CCCCCC"),
                                top=Side(style="thin", color="CCCCCC"),
                                bottom=Side(style="thin", color="CCCCCC")
                            )

                            for col_num, column_title in enumerate(existing_cols, start=1):
                                cell = ws3.cell(row=1, column=col_num, value=column_title)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = align_center
                                cell.border = border

                            for row_num, row_data in enumerate(final_siwa[existing_cols].values, start=2):
                                for col_num, cell_value in enumerate(row_data, start=1):
                                    cell = ws3.cell(row=row_num, column=col_num, value=str(cell_value))
                                    if row_num % 2 == 0:
                                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                    cell.alignment = align_center
                                    cell.border = border

                            for col in ws3.columns:
                                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                                ws3.column_dimensions[col[0].column_letter].width = max_length + 2

                            footer_row = len(final_siwa) + 3
                            ws3.merge_cells(f"A{footer_row}:{ws3.cell(1, len(existing_cols)).column_letter}{footer_row}")
                            ws3.cell(row=footer_row, column=1).value = f"Generated by GWA MASR | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                            ws3.cell(row=footer_row, column=1).alignment = align_center
                            ws3.cell(row=footer_row, column=1).font = Font(color="888888", italic=True, size=10)

                            buffer3 = BytesIO()
                            wb3.save(buffer3)
                            buffer3.seek(0)

                            st.download_button(
                                label="📘 Download Situation Siwa (Excel)",
                                data=buffer3,
                                file_name=f"Situation_Siwa_{selected_date}_{selected_from}_{selected_nights}_nights.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

    with tab4:
        st.markdown("## 🛏️ Rooming List")

        # محاولة تحديد الأعمدة تلقائيًا
        checkin_col = next((c for c in df.columns if "check" in c.lower() and "in" in c.lower()), None)
        hotel_col = next((c for c in df.columns if "hotel" in c.lower()), None)
        name_col = next((c for c in df.columns if "name" in c.lower()), None)
        rooms_col = next((c for c in df.columns if "room" in c.lower()), None)

        # التحقق من الأعمدة
        if not all([checkin_col, hotel_col, name_col, rooms_col]):
            st.error("❌ One or more required columns not found in your sheet.")
        else:
            # تحويل التاريخ لتاريخ فقط بدون وقت
            df[checkin_col] = pd.to_datetime(df[checkin_col], errors='coerce').dt.date

            # ✅ فلتر التاريخ
            check_in_dates = sorted(df[checkin_col].dropna().unique())
            selected_date = st.selectbox("📅 Select Check In Date", check_in_dates, key="rooming_date")

            if selected_date:
                filtered_by_date = df[df[checkin_col] == selected_date]

                # ✅ فلتر الفندق
                available_hotels = sorted(filtered_by_date[hotel_col].dropna().unique())
                selected_hotel = st.selectbox("🏨 Select Hotel", available_hotels, key="rooming_hotel")

                if selected_hotel:
                    rooming_df = filtered_by_date[filtered_by_date[hotel_col] == selected_hotel]

                    if rooming_df.empty:
                        st.warning("⚠️ No guests found for this hotel and date.")
                    else:
                        st.success(f"✅ Found {len(rooming_df)} guests staying at {selected_hotel} on {selected_date}")

                        # ✅ عرض فقط العمودين المطلوبين
                        display_cols = [name_col, rooms_col]
                        st.dataframe(rooming_df[display_cols], use_container_width=True)

                        # 🔷 إنشاء ملف Excel بنفس التنسيق الأزرق
                        wb4 = Workbook()
                        ws4 = wb4.active
                        ws4.title = "Rooming List"

                        header_fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        align_center = Alignment(horizontal="center", vertical="center")
                        border = Border(
                            left=Side(style="thin", color="CCCCCC"),
                            right=Side(style="thin", color="CCCCCC"),
                            top=Side(style="thin", color="CCCCCC"),
                            bottom=Side(style="thin", color="CCCCCC")
                        )

                        # العناوين
                        for col_num, column_title in enumerate(display_cols, start=1):
                            cell = ws4.cell(row=1, column=col_num, value=column_title)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = align_center
                            cell.border = border

                        # البيانات
                        for row_num, row_data in enumerate(rooming_df[display_cols].values, start=2):
                            for col_num, cell_value in enumerate(row_data, start=1):
                                cell = ws4.cell(row=row_num, column=col_num, value=str(cell_value))
                                if row_num % 2 == 0:
                                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                cell.alignment = align_center
                                cell.border = border

                        # عرض الأعمدة المناسب
                        for col in ws4.columns:
                            max_length = max(len(str(cell.value)) for cell in col if cell.value)
                            ws4.column_dimensions[col[0].column_letter].width = max_length + 2

                        # تذييل احترافي
                        footer_row = len(rooming_df) + 3
                        ws4.merge_cells(f"A{footer_row}:{ws4.cell(1, len(display_cols)).column_letter}{footer_row}")
                        ws4.cell(row=footer_row, column=1).value = f"Generated by GWA MASR | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                        ws4.cell(row=footer_row, column=1).alignment = align_center
                        ws4.cell(row=footer_row, column=1).font = Font(color="888888", italic=True, size=10)

                        # زر التحميل
                        buffer4 = BytesIO()
                        wb4.save(buffer4)
                        buffer4.seek(0)

                        st.download_button(
                            label="📘 Download Rooming List (Excel)",
                            data=buffer4,
                            file_name=f"Rooming_List_{selected_hotel}_{selected_date}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

    with tab5:
        st.markdown("## 🏨 Hotel Room Availability — Daily Snapshot")

        # اختيار التاريخ من التواريخ الموجودة في الشيت
        available_dates = sorted(df[checkin_col].dropna().unique())
        selected_date_avail = st.selectbox("📅 اختر تاريخ لعرض Availability", available_dates)

        if selected_date_avail:
            avail_df = df[df[checkin_col] == selected_date_avail].copy()
            
            if avail_df.empty:
                st.warning("⚠️ لا توجد حجوزات في التاريخ المحدد.")
            else:
                # اكتشاف عمود نوع الغرفة والفندق
                room_type_col = next((c for c in df.columns if any(w in c.lower() for w in ["room", "type", "category"])), None)
                hotel_col = next((c for c in df.columns if any(w in c.lower() for w in ["hotel", "resort", "place"])), None)

                if not hotel_col or not room_type_col:
                    st.error("❌ لم أجد عمود الفندق أو نوع الغرفة في الملف.")
                else:
                    # تنظيف عمود الفندق
                    avail_df[hotel_col] = avail_df[hotel_col].astype(str).str.strip().str.title()

                    # تجميع حسب الفندق
                    hotel_group = avail_df.groupby(hotel_col, dropna=False)
                    
                    cols_per_row = 3
                    hotels_list = list(hotel_group.groups.keys())
                    rows = (len(hotels_list) + cols_per_row - 1) // cols_per_row
                    idx = 0

                    for r in range(rows):
                        cols = st.columns(cols_per_row)
                        for c in cols:
                            if idx >= len(hotels_list):
                                idx += 1
                                continue

                            hotel = hotels_list[idx]
                            hotel_data = hotel_group.get_group(hotel)
                            idx += 1

                            # عدد الغرف لكل نوع
                            room_counts = hotel_data[room_type_col].value_counts().to_dict()
                            room_summary = '<br>'.join([f"{v} {k}" for k, v in room_counts.items()])

                            # HTML box للفندق
                            # HTML box للفندق بلون مشابه للـ Home Overview
                            box_html = f"""
                            <div style="
                                border-radius:12px;
                                padding:14px;
                                margin:8px;
                                background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(0,0,0,0.06));
                                border:1px solid rgba(255,255,255,0.06);
                                color:white;
                                font-weight:bold;
                            ">
                                <h4 style="margin:0; color:#FFD700;">🏨 {hotel}</h4>
                                <p style="margin:6px 0 4px 0; line-height:1.4;">{room_summary}</p>
                                <p style="margin:4px 0 0 0;"><b>Total Rooms:</b> {hotel_data.shape[0]}</p>
                            </div>
                                """

                            c.markdown(box_html, unsafe_allow_html=True)

                            # Expander للتفاصيل
                            with c.expander(f"🔎 View bookings for {hotel}"):
                                st.dataframe(hotel_data.reset_index(drop=True), use_container_width=True)





except Exception as e:
    st.error(f"❌ Error loading sheet: {e}")
